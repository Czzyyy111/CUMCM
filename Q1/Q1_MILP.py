"""2024 年高教社杯数学建模竞赛 C 题问题一：确定性 MILP。

本脚本同时求解：
1. 超过预期销售量的部分滞销并浪费；
2. 超过预期销售量的部分按原价 50% 销售。

模型以地块—作物—年份—季次的种植面积为核心变量，包含土地适配、
水浇地种植模式、连续重茬、滚动三年豆类轮作、最小种植面积和分散度约束。
原始附件只读，所有结果写入 Q1 目录。
"""

from __future__ import annotations

import json
import math
import shutil
import time
import warnings
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pulp
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# =============================================================================
# 全局配置（题面未量化的管理规则集中在此，便于敏感性分析）
# =============================================================================

RANDOM_SEED = 2024
YEARS = list(range(2024, 2031))
SEASONS = (1, 2)
BEAN_CROPS = set(range(1, 6)) | set(range(17, 20))
GRAIN_CROPS = set(range(1, 16))
VEG_FIRST_CROPS = set(range(17, 35))
VEG_SECOND_D_CROPS = set(range(35, 38))
FUNGI_CROPS = set(range(38, 42))

# “面积不宜太小”和“种植地不能太分散”的可审计量化假设。
MIN_AREA_OPEN = 5.0       # 露天地块中单种作物的最小种植面积（亩）
MIN_AREA_GREENHOUSE = 0.3 # 大棚中单种作物的最小种植面积（亩）
MAX_PLOTS_PER_CROP = 7    # 同一作物在同一年同一季最多分布的地块数

SOLVER_TIME_LIMIT = 240   # 每种情形 HiGHS 最长求解时间（秒）
SOLVER_GAP = 0.01         # 相对最优间隙；达到 1% 即视为竞赛可接受近似最优
EPS = 1e-6

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
DATA_DIR = ROOT_DIR / "C题"
OUTPUT_DIR = SCRIPT_DIR / "outputs"
FIGURE_DIR = SCRIPT_DIR / "figures"
LOG_DIR = SCRIPT_DIR / "logs"

ATTACHMENT_1 = DATA_DIR / "附件1.xlsx"
ATTACHMENT_2 = DATA_DIR / "附件2.xlsx"
TEMPLATE_1 = DATA_DIR / "附件3" / "result1_1.xlsx"
TEMPLATE_2 = DATA_DIR / "附件3" / "result1_2.xlsx"


@dataclass(frozen=True)
class Plot:
    name: str
    land_type: str
    area: float


@dataclass(frozen=True)
class Stat:
    crop_id: int
    crop_name: str
    land_type: str
    season_label: str
    yield_per_mu: float
    cost_per_mu: float
    price: float


@dataclass
class InputData:
    plots: dict[str, Plot]
    crop_names: dict[int, str]
    crop_types: dict[int, str]
    stats: dict[tuple[int, str, str], Stat]
    planting_2023: list[dict]
    sales_demand: dict[int, float]


@dataclass
class ScenarioResult:
    scenario: str
    discount_rate: float
    status: str
    solution_status: str
    objective: float
    elapsed_seconds: float
    x_values: dict[tuple[str, int, int, int], float]
    y_values: dict[tuple[str, int, int, int], int]
    normal_sales: dict[tuple[int, int, float], float]
    annual_summary: pd.DataFrame
    crop_summary: pd.DataFrame
    land_summary: pd.DataFrame
    diagnostics: dict


def clean_text(value: object) -> str:
    """清理 Excel 中可能出现的首尾空格和换行。"""
    return "" if value is None else str(value).strip()


def parse_price(value: object) -> float:
    """将价格区间转换为中点；单值价格直接转为浮点数。"""
    text = clean_text(value).replace("—", "-").replace("–", "-")
    if "-" in text:
        low, high = text.split("-", maxsplit=1)
        return (float(low) + float(high)) / 2.0
    return float(text)


def season_label(raw: object) -> str:
    text = clean_text(raw)
    if "单季" in text:
        return "单季"
    if "第一" in text:
        return "第一季"
    if "第二" in text:
        return "第二季"
    raise ValueError(f"无法识别季次：{raw!r}")


def configure_chinese_plot() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Arial Unicode MS"
    ]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 120
    plt.rcParams["savefig.dpi"] = 300


# =============================================================================
# STAGE A 读取并校验数据
# =============================================================================

def load_input_data() -> InputData:
    for path in (ATTACHMENT_1, ATTACHMENT_2, TEMPLATE_1, TEMPLATE_2):
        if not path.exists():
            raise FileNotFoundError(f"缺少输入文件：{path}")

    wb1 = load_workbook(ATTACHMENT_1, data_only=True, read_only=True)
    ws_plot = wb1["乡村的现有耕地"]
    plots: dict[str, Plot] = {}
    for row in ws_plot.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = clean_text(row[0])
        plots[name] = Plot(name, clean_text(row[1]), float(row[2]))

    ws_crop = wb1["乡村种植的农作物"]
    crop_names: dict[int, str] = {}
    crop_types: dict[int, str] = {}
    for row in ws_crop.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        crop_id = int(row[0])
        crop_names[crop_id] = clean_text(row[1])
        crop_types[crop_id] = clean_text(row[2])
    wb1.close()

    wb2 = load_workbook(ATTACHMENT_2, data_only=True, read_only=True)
    ws_2023 = wb2["2023年的农作物种植情况"]
    planting_2023: list[dict] = []
    current_plot = ""
    for row in ws_2023.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            current_plot = clean_text(row[0])
        if not current_plot or not isinstance(row[1], (int, float)):
            continue
        planting_2023.append({
            "plot": current_plot,
            "crop_id": int(row[1]),
            "crop_name": clean_text(row[2]),
            "area": float(row[4]),
            "season": season_label(row[5]),
        })

    ws_stats = wb2["2023年统计的相关数据"]
    stats: dict[tuple[int, str, str], Stat] = {}
    for row in ws_stats.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        crop_id = int(row[1])
        land_type = clean_text(row[3])
        s_label = season_label(row[4])
        stat = Stat(
            crop_id=crop_id,
            crop_name=clean_text(row[2]),
            land_type=land_type,
            season_label=s_label,
            yield_per_mu=float(row[5]),
            cost_per_mu=float(row[6]),
            price=parse_price(row[7]),
        )
        stats[(crop_id, land_type, s_label)] = stat
    wb2.close()

    if len(plots) != 54 or len(crop_names) != 41:
        raise ValueError(f"基础数据规模异常：地块 {len(plots)} 个，作物 {len(crop_names)} 种")

    data = InputData(plots, crop_names, crop_types, stats, planting_2023, {})
    demand = defaultdict(float)
    for record in planting_2023:
        plot = plots[record["plot"]]
        s_num = 1 if record["season"] in {"单季", "第一季"} else 2
        stat = get_stat(data, plot, record["crop_id"], s_num)
        demand[record["crop_id"]] += record["area"] * stat.yield_per_mu
    data.sales_demand = {crop: float(demand.get(crop, 0.0)) for crop in crop_names}

    missing_demand = [crop for crop, value in data.sales_demand.items() if value <= 0]
    if missing_demand:
        raise ValueError(f"以下作物无法由 2023 年种植数据得到销售量基准：{missing_demand}")
    return data


def get_stat(data: InputData, plot: Plot, crop_id: int, season: int) -> Stat:
    """按地块、作物和季次取得亩产、成本、价格参数。"""
    if plot.land_type in {"平旱地", "梯田", "山坡地"}:
        key = (crop_id, plot.land_type, "单季")
    elif plot.land_type == "水浇地":
        key = (crop_id, "水浇地", "单季" if crop_id == 16 else f"第{'一' if season == 1 else '二'}季")
    elif plot.land_type == "普通大棚":
        key = (crop_id, "普通大棚", f"第{'一' if season == 1 else '二'}季")
    elif plot.land_type == "智慧大棚":
        # 附件 2 注明：智慧大棚第一季与普通大棚第一季参数相同。
        key = (crop_id, "普通大棚", "第一季") if season == 1 else (crop_id, "智慧大棚", "第二季")
    else:
        raise KeyError(f"未知地块类型：{plot.land_type}")
    if key not in data.stats:
        raise KeyError(f"缺少统计参数：{key}，地块={plot.name}，季次={season}")
    return data.stats[key]


def allowed_crops(plot: Plot, season: int) -> list[int]:
    if plot.land_type in {"平旱地", "梯田", "山坡地"}:
        return sorted(GRAIN_CROPS) if season == 1 else []
    if plot.land_type == "水浇地":
        return sorted({16} | VEG_FIRST_CROPS) if season == 1 else sorted(VEG_SECOND_D_CROPS)
    if plot.land_type == "普通大棚":
        return sorted(VEG_FIRST_CROPS if season == 1 else FUNGI_CROPS)
    if plot.land_type == "智慧大棚":
        return sorted(VEG_FIRST_CROPS)
    return []


def minimum_area(plot: Plot) -> float:
    return MIN_AREA_GREENHOUSE if "大棚" in plot.land_type else min(MIN_AREA_OPEN, plot.area)


def fixed_2023_bean_area(data: InputData, plot_name: str) -> float:
    return sum(
        row["area"] for row in data.planting_2023
        if row["plot"] == plot_name and row["crop_id"] in BEAN_CROPS
    )


# =============================================================================
# STAGE B 构造 MILP
# =============================================================================

def build_and_solve(data: InputData, scenario: str, discount_rate: float) -> ScenarioResult:
    start_time = time.perf_counter()
    model = pulp.LpProblem(f"Q1_{scenario}", pulp.LpMaximize)

    x: dict[tuple[str, int, int, int], pulp.LpVariable] = {}
    y: dict[tuple[str, int, int, int], pulp.LpVariable] = {}
    stat_for_key: dict[tuple[str, int, int, int], Stat] = {}

    for plot in data.plots.values():
        for year in YEARS:
            for season in SEASONS:
                for crop in allowed_crops(plot, season):
                    key = (plot.name, year, season, crop)
                    x[key] = pulp.LpVariable(f"x_{plot.name}_{year}_{season}_{crop}", lowBound=0)
                    y[key] = pulp.LpVariable(f"y_{plot.name}_{year}_{season}_{crop}", cat="Binary")
                    stat_for_key[key] = get_stat(data, plot, crop, season)
                    model += x[key] <= plot.area * y[key], f"面积上界_{plot.name}_{year}_{season}_{crop}"
                    model += x[key] >= minimum_area(plot) * y[key], f"最小面积_{plot.name}_{year}_{season}_{crop}"

    # 地块面积及水浇地种植模式约束。
    water_rice_mode: dict[tuple[str, int], pulp.LpVariable] = {}
    for plot in data.plots.values():
        for year in YEARS:
            if plot.land_type == "水浇地":
                z = pulp.LpVariable(f"水稻模式_{plot.name}_{year}", cat="Binary")
                water_rice_mode[(plot.name, year)] = z
                model += x[(plot.name, year, 1, 16)] == plot.area * z
                model += pulp.lpSum(x[(plot.name, year, 1, c)] for c in VEG_FIRST_CROPS) == plot.area * (1 - z)
                model += pulp.lpSum(x[(plot.name, year, 2, c)] for c in VEG_SECOND_D_CROPS) == plot.area * (1 - z)
            else:
                for season in SEASONS:
                    crops = allowed_crops(plot, season)
                    if crops:
                        model += pulp.lpSum(x[(plot.name, year, season, c)] for c in crops) == plot.area

    # 同一作物在同一地块不得连续重茬。
    # 单季露天地与水稻跨年检查；智慧大棚按真实种植季次顺序检查。
    actual_2023 = defaultdict(set)
    for row in data.planting_2023:
        s_num = 1 if row["season"] in {"单季", "第一季"} else 2
        actual_2023[(row["plot"], s_num)].add(row["crop_id"])

    for plot in data.plots.values():
        if plot.land_type in {"平旱地", "梯田", "山坡地"}:
            for crop in GRAIN_CROPS:
                if crop in actual_2023[(plot.name, 1)]:
                    model += y[(plot.name, 2024, 1, crop)] == 0
                for year in YEARS[:-1]:
                    model += y[(plot.name, year, 1, crop)] + y[(plot.name, year + 1, 1, crop)] <= 1
        elif plot.land_type == "水浇地":
            if 16 in actual_2023[(plot.name, 1)]:
                model += y[(plot.name, 2024, 1, 16)] == 0
            for year in YEARS[:-1]:
                model += y[(plot.name, year, 1, 16)] + y[(plot.name, year + 1, 1, 16)] <= 1
        elif plot.land_type == "智慧大棚":
            for crop in VEG_FIRST_CROPS:
                if crop in actual_2023[(plot.name, 2)]:
                    model += y[(plot.name, 2024, 1, crop)] == 0
                for year in YEARS:
                    model += y[(plot.name, year, 1, crop)] + y[(plot.name, year, 2, crop)] <= 1
                for year in YEARS[:-1]:
                    model += y[(plot.name, year, 2, crop)] + y[(plot.name, year + 1, 1, crop)] <= 1

    # 从 2023 年开始，每个滚动三年窗口内每块地累计豆类面积至少等于地块面积。
    for plot in data.plots.values():
        fixed_2023 = fixed_2023_bean_area(data, plot.name)
        for window_start in range(2023, 2029):
            terms = []
            constant = fixed_2023 if window_start == 2023 else 0.0
            for year in range(max(2024, window_start), window_start + 3):
                for season in SEASONS:
                    for crop in BEAN_CROPS:
                        key = (plot.name, year, season, crop)
                        if key in x:
                            terms.append(x[key])
            model += pulp.lpSum(terms) + constant >= plot.area, f"豆类轮作_{plot.name}_{window_start}"

    # 每种作物每年每季至多分布在指定数量的地块，避免过度分散。
    for year in YEARS:
        for season in SEASONS:
            for crop in data.crop_names:
                keys = [key for key in y if key[1] == year and key[2] == season and key[3] == crop]
                if keys:
                    model += pulp.lpSum(y[key] for key in keys) <= MAX_PLOTS_PER_CROP

    # 按“作物—年份—销售价格组”汇总产量。智慧大棚第二季价格可能更高，
    # 因而允许正常销量优先分配给高价格批次。
    production_terms: dict[tuple[int, int, float], list] = defaultdict(list)
    for key, var in x.items():
        _, year, _, crop = key
        stat = stat_for_key[key]
        production_terms[(crop, year, stat.price)].append(stat.yield_per_mu * var)

    normal_sales: dict[tuple[int, int, float], pulp.LpVariable] = {}
    production_expr: dict[tuple[int, int, float], pulp.LpAffineExpression] = {}
    for group, terms in production_terms.items():
        crop, year, price = group
        expr = pulp.lpSum(terms)
        production_expr[group] = expr
        sale = pulp.LpVariable(f"正常销量_{crop}_{year}_{price:.4f}", lowBound=0)
        normal_sales[group] = sale
        model += sale <= expr

    for crop in data.crop_names:
        for year in YEARS:
            groups = [group for group in normal_sales if group[0] == crop and group[1] == year]
            model += pulp.lpSum(normal_sales[g] for g in groups) <= data.sales_demand[crop]

    revenue = pulp.lpSum(
        discount_rate * group[2] * production_expr[group]
        + (1.0 - discount_rate) * group[2] * normal_sales[group]
        for group in production_expr
    )
    cost = pulp.lpSum(stat_for_key[key].cost_per_mu * var for key, var in x.items())
    model += revenue - cost

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    solver_log = LOG_DIR / f"HiGHS_{scenario}.log"
    solver = pulp.HiGHS(
        msg=True,
        timeLimit=SOLVER_TIME_LIMIT,
        gapRel=SOLVER_GAP,
        threads=4,
        log_file=str(solver_log),
    )
    if not solver.available():
        raise RuntimeError("PuLP 未找到 HiGHS 求解器，请重新安装 requirements.txt 中的 highspy。")
    model.solve(solver)
    elapsed = time.perf_counter() - start_time
    pulp_status = pulp.LpStatus.get(model.status, str(model.status))
    solution_status = pulp.LpSolution.get(model.sol_status, str(model.sol_status))
    highs_model_status = model.solverModel.modelStatusToString(model.solverModel.getModelStatus())
    highs_info = model.solverModel.getInfo()
    mip_gap = float(highs_info.mip_gap)
    dual_bound = -float(highs_info.mip_dual_bound)  # PuLP 最大化问题在 HiGHS 内部按最小化处理
    if "Time limit" in highs_model_status:
        status = "达到时间上限，已获得可行整数解"
    elif "Optimal" in highs_model_status:
        status = f"达到设定最优间隙（{mip_gap:.3%}）"
    else:
        status = highs_model_status
    if model.sol_status not in {pulp.LpSolutionOptimal, pulp.LpSolutionIntegerFeasible}:
        raise RuntimeError(f"情形 {scenario} 未得到可行整数解：status={status}, solution={solution_status}")

    x_values = {key: max(0.0, float(pulp.value(var) or 0.0)) for key, var in x.items()}
    y_values = {key: int(round(float(pulp.value(var) or 0.0))) for key, var in y.items()}
    sale_values = {group: max(0.0, float(pulp.value(var) or 0.0)) for group, var in normal_sales.items()}

    annual, crop_summary, land_summary = summarize_solution(
        data, x_values, sale_values, stat_for_key, discount_rate
    )
    diagnostics = validate_solution(data, x_values, y_values, annual)
    diagnostics.update({
        "模型名称": "确定性多期混合整数线性规划（MILP）",
        "情形": scenario,
        "超产销售折扣率": discount_rate,
        "求解状态": status,
        "PuLP状态映射": pulp_status,
        "整数解状态": solution_status,
        "HiGHS原始状态": highs_model_status,
        "最终相对间隙": mip_gap,
        "目标函数理论上界_元": dual_bound,
        "目标函数值_元": float(pulp.value(model.objective)),
        "求解耗时_秒": elapsed,
        "变量总数": len(model.variables()),
        "二进制变量数": len(y) + len(water_rice_mode),
        "约束总数": len(model.constraints),
        "HiGHS相对间隙设置": SOLVER_GAP,
        "HiGHS时间上限_秒": SOLVER_TIME_LIMIT,
    })
    return ScenarioResult(
        scenario=scenario,
        discount_rate=discount_rate,
        status=status,
        solution_status=solution_status,
        objective=float(pulp.value(model.objective)),
        elapsed_seconds=elapsed,
        x_values=x_values,
        y_values=y_values,
        normal_sales=sale_values,
        annual_summary=annual,
        crop_summary=crop_summary,
        land_summary=land_summary,
        diagnostics=diagnostics,
    )


# =============================================================================
# STAGE C 结果汇总与约束诊断
# =============================================================================

def summarize_solution(
    data: InputData,
    x_values: dict[tuple[str, int, int, int], float],
    sales_values: dict[tuple[int, int, float], float],
    stat_for_key: dict[tuple[str, int, int, int], Stat],
    discount_rate: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = []
    land_rows = []
    for year in YEARS:
        for crop, crop_name in data.crop_names.items():
            crop_keys = [key for key in x_values if key[1] == year and key[3] == crop]
            area = sum(x_values[key] for key in crop_keys)
            production = sum(x_values[key] * stat_for_key[key].yield_per_mu for key in crop_keys)
            cost = sum(x_values[key] * stat_for_key[key].cost_per_mu for key in crop_keys)
            group_sales = {
                group: value for group, value in sales_values.items()
                if group[0] == crop and group[1] == year
            }
            normal = sum(group_sales.values())
            revenue = 0.0
            for price in sorted({stat_for_key[key].price for key in crop_keys}):
                group = (crop, year, price)
                group_prod = sum(
                    x_values[key] * stat_for_key[key].yield_per_mu
                    for key in crop_keys if math.isclose(stat_for_key[key].price, price)
                )
                group_normal = group_sales.get(group, 0.0)
                revenue += price * group_normal + discount_rate * price * max(0.0, group_prod - group_normal)
            rows.append({
                "年份": year,
                "作物编号": crop,
                "作物名称": crop_name,
                "作物类型": data.crop_types[crop],
                "种植面积/亩": area,
                "总产量/斤": production,
                "预期销售量/斤": data.sales_demand[crop],
                "正常销量/斤": normal,
                "超出销量/斤": max(0.0, production - normal),
                "销售收入/元": revenue,
                "种植成本/元": cost,
                "利润/元": revenue - cost,
            })

        for land_type in sorted({p.land_type for p in data.plots.values()}):
            for season in SEASONS:
                area = sum(
                    value for key, value in x_values.items()
                    if key[1] == year and key[2] == season
                    and data.plots[key[0]].land_type == land_type
                )
                land_rows.append({"年份": year, "地块类型": land_type, "季次": f"第{season}季", "种植面积/亩": area})

    crop_df = pd.DataFrame(rows)
    annual = crop_df.groupby("年份", as_index=False).agg({
        "种植面积/亩": "sum",
        "总产量/斤": "sum",
        "正常销量/斤": "sum",
        "超出销量/斤": "sum",
        "销售收入/元": "sum",
        "种植成本/元": "sum",
        "利润/元": "sum",
    })
    return annual, crop_df, pd.DataFrame(land_rows)


def validate_solution(
    data: InputData,
    x_values: dict[tuple[str, int, int, int], float],
    y_values: dict[tuple[str, int, int, int], int],
    annual: pd.DataFrame,
) -> dict:
    max_land_error = 0.0
    min_area_violations = 0
    dispersion_violations = 0
    bean_violations = 0
    rotation_violations = 0

    for key, value in x_values.items():
        plot = data.plots[key[0]]
        indicator = y_values[key]
        if value > EPS and value + 1e-5 < minimum_area(plot):
            min_area_violations += 1
        if indicator == 0 and value > 1e-5:
            min_area_violations += 1

    for plot in data.plots.values():
        for year in YEARS:
            if plot.land_type == "水浇地":
                rice = x_values[(plot.name, year, 1, 16)]
                veg1 = sum(x_values[(plot.name, year, 1, c)] for c in VEG_FIRST_CROPS)
                veg2 = sum(x_values[(plot.name, year, 2, c)] for c in VEG_SECOND_D_CROPS)
                if rice > plot.area / 2:
                    max_land_error = max(max_land_error, abs(rice - plot.area), veg1, veg2)
                else:
                    max_land_error = max(max_land_error, rice, abs(veg1 - plot.area), abs(veg2 - plot.area))
            else:
                for season in SEASONS:
                    crops = allowed_crops(plot, season)
                    if crops:
                        used = sum(x_values[(plot.name, year, season, c)] for c in crops)
                        max_land_error = max(max_land_error, abs(used - plot.area))

        fixed = fixed_2023_bean_area(data, plot.name)
        for start in range(2023, 2029):
            bean_area = fixed if start == 2023 else 0.0
            for year in range(max(2024, start), start + 3):
                bean_area += sum(
                    value for key, value in x_values.items()
                    if key[0] == plot.name and key[1] == year and key[3] in BEAN_CROPS
                )
            if bean_area + 1e-5 < plot.area:
                bean_violations += 1

    for year in YEARS:
        for season in SEASONS:
            for crop in data.crop_names:
                count = sum(
                    indicator for key, indicator in y_values.items()
                    if key[1] == year and key[2] == season and key[3] == crop
                )
                if count > MAX_PLOTS_PER_CROP:
                    dispersion_violations += 1

    # 按模型采用的重茬口径复核未来相邻种植事件。
    for plot in data.plots.values():
        if plot.land_type in {"平旱地", "梯田", "山坡地"}:
            for crop in GRAIN_CROPS:
                for year in YEARS[:-1]:
                    rotation_violations += int(
                        y_values[(plot.name, year, 1, crop)] + y_values[(plot.name, year + 1, 1, crop)] > 1
                    )
        elif plot.land_type == "水浇地":
            for year in YEARS[:-1]:
                rotation_violations += int(
                    y_values[(plot.name, year, 1, 16)] + y_values[(plot.name, year + 1, 1, 16)] > 1
                )
        elif plot.land_type == "智慧大棚":
            for crop in VEG_FIRST_CROPS:
                for year in YEARS:
                    rotation_violations += int(
                        y_values[(plot.name, year, 1, crop)] + y_values[(plot.name, year, 2, crop)] > 1
                    )
                for year in YEARS[:-1]:
                    rotation_violations += int(
                        y_values[(plot.name, year, 2, crop)] + y_values[(plot.name, year + 1, 1, crop)] > 1
                    )

    return {
        "原始地块数": len(data.plots),
        "原始作物数": len(data.crop_names),
        "2023种植记录数": len(data.planting_2023),
        "规划年份数": len(YEARS),
        "面积平衡最大误差_亩": max_land_error,
        "最小面积约束违规数": min_area_violations,
        "分散度约束违规数": dispersion_violations,
        "豆类轮作约束违规数": bean_violations,
        "重茬约束违规数": rotation_violations,
        "年度利润是否均为有限数": bool(np.isfinite(annual["利润/元"]).all()),
        "数据泄漏检查": "不适用：本问为确定性优化；2024—2030 未使用任何未来观测值",
        "销售量口径": "2023年各作物实际种植面积×对应亩产量，按作物汇总",
        "销售价格口径": "附件价格区间中点；智慧大棚第二季使用附件单列价格",
        "最小种植面积假设": f"露天地块 {MIN_AREA_OPEN} 亩，大棚 {MIN_AREA_GREENHOUSE} 亩",
        "最大分布地块数假设": f"每种作物每年每季最多 {MAX_PLOTS_PER_CROP} 个地块",
        "随机种子": RANDOM_SEED,
    }


# =============================================================================
# STAGE D 写入竞赛模板和分析工作簿
# =============================================================================

def write_template(data: InputData, result: ScenarioResult, template: Path, output: Path) -> None:
    shutil.copy2(template, output)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(output)
    for year in YEARS:
        ws = wb[str(year)]
        crop_col = {
            clean_text(ws.cell(1, col).value): col
            for col in range(3, ws.max_column + 1)
            if ws.cell(1, col).value is not None
        }
        # 模板第一季：第 2—55 行；第二季：第 56—83 行。
        for row_start, row_end, season in ((2, 55, 1), (56, 83, 2)):
            for row in range(row_start, row_end + 1):
                plot_name = clean_text(ws.cell(row, 2).value)
                if not plot_name:
                    continue
                for crop, crop_name in data.crop_names.items():
                    col = crop_col[crop_name]
                    value = result.x_values.get((plot_name, year, season, crop), 0.0)
                    ws.cell(row, col).value = 0 if value < 5e-5 else round(value, 4)
    wb.save(output)


def style_dataframe_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2, 24)
        ws.column_dimensions[col[0].column_letter].width = max(width, 10)


def append_dataframe(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([float(v) if isinstance(v, np.floating) else int(v) if isinstance(v, np.integer) else v for v in row])
    style_dataframe_sheet(ws)


def write_analysis_workbook(data: InputData, results: list[ScenarioResult], output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    comparison = pd.concat([
        result.annual_summary.assign(情形=result.scenario)
        for result in results
    ], ignore_index=True)
    comparison = comparison[["情形", "年份", "种植面积/亩", "总产量/斤", "正常销量/斤", "超出销量/斤", "销售收入/元", "种植成本/元", "利润/元"]]
    append_dataframe(wb, "年度汇总", comparison)

    for result in results:
        short = "滞销浪费" if result.discount_rate == 0 else "半价销售"
        append_dataframe(wb, f"作物明细_{short}", result.crop_summary)
        append_dataframe(wb, f"土地利用_{short}", result.land_summary)

    demand_df = pd.DataFrame([
        {"作物编号": crop, "作物名称": data.crop_names[crop], "2023预期销售量基准/斤": data.sales_demand[crop]}
        for crop in data.crop_names
    ])
    append_dataframe(wb, "销售量基准", demand_df)

    method_rows = [
        ("模型", "2024—2030 年确定性多期混合整数线性规划（MILP）"),
        ("决策变量", "种植面积 x；是否种植 y；水浇地水稻模式 z；各价格组正常销量 q"),
        ("情形1利润", "正常销售收入－种植成本；超出预期销售量部分收入为0"),
        ("情形2利润", "正常销售收入＋0.5×超产部分收入－种植成本"),
        ("预期销售量", "由2023年实际种植面积乘对应亩产量后按作物汇总"),
        ("销售价格", "采用附件2价格区间中点；智慧大棚第二季使用其单列价格"),
        ("土地约束", "按平旱地、梯田、山坡地、水浇地、普通大棚、智慧大棚的适种规则建模"),
        ("重茬约束", "单季地与水稻跨年不连续；智慧大棚按第一季—第二季—次年第一季检查"),
        ("豆类轮作", "每个滚动三年窗口内，每个地块累计豆类面积不少于该地块面积"),
        ("管理便利", f"露天地块最小{MIN_AREA_OPEN}亩，大棚最小{MIN_AREA_GREENHOUSE}亩；每作物每年每季最多{MAX_PLOTS_PER_CROP}个地块"),
        ("求解器", "PuLP + HiGHS，开源且无需商业许可证"),
        ("重要限制", "最小面积和最大地块数是对题目定性要求的量化假设，论文中应进行敏感性分析"),
    ]
    append_dataframe(wb, "方法说明", pd.DataFrame(method_rows, columns=["项目", "说明"]))

    diag_rows = []
    for result in results:
        for key, value in result.diagnostics.items():
            diag_rows.append({"情形": result.scenario, "检查项": key, "结果": value})
    append_dataframe(wb, "诊断日志", pd.DataFrame(diag_rows))
    wb.save(output)


def write_text_log(results: list[ScenarioResult], output: Path) -> None:
    lines = [
        "问题一 MILP 运行诊断日志",
        "=" * 60,
        "",
        "模型口径：确定性多期混合整数线性规划。",
        "2023 年实际产量汇总作为各作物预期销售量基准。",
        "所有未来年度参数相对 2023 年保持不变。",
        "",
    ]
    for result in results:
        lines.append(f"【{result.scenario}】")
        for key, value in result.diagnostics.items():
            lines.append(f"{key}: {value}")
        lines.append("")
    lines.extend([
        "已执行的有效性检查：",
        "1. 地块和季次面积平衡；",
        "2. 最小种植面积及二进制联动；",
        "3. 每作物每季最大分布地块数；",
        "4. 滚动三年豆类轮作；",
        "5. 连续重茬限制；",
        "6. 年度收入、成本和利润有限性。",
        "",
        "已知限制：题面未给出管理便利约束的具体阈值，当前采用的阈值是工程化假设，",
        "建议后续围绕最小种植面积和最大分布地块数开展敏感性分析。",
    ])
    output.write_text("\n".join(lines), encoding="utf-8")


# =============================================================================
# STAGE E 中文结果图
# =============================================================================

def crop_category(name: str) -> str:
    if "豆类" in name:
        return "豆类"
    if "食用菌" in name:
        return "食用菌"
    if "蔬菜" in name:
        return "其他蔬菜"
    return "其他粮食"


def create_figures(results: list[ScenarioResult]) -> None:
    configure_chinese_plot()
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    colors = ["#2878B5", "#C82423"]

    # 图1：年度利润比较。
    fig, ax = plt.subplots(figsize=(9, 5.2))
    width = 0.36
    positions = np.arange(len(YEARS))
    for idx, result in enumerate(results):
        profit = result.annual_summary["利润/元"].to_numpy() / 10000
        offset = (idx - 0.5) * width
        bars = ax.bar(positions + offset, profit, width, label=result.scenario, color=colors[idx], alpha=0.88)
        ax.bar_label(bars, fmt="%.1f", fontsize=8, padding=2)
    ax.set_title("两种滞销处理情形下的年度利润比较")
    ax.set_xlabel("年份")
    ax.set_ylabel("利润（万元）")
    ax.set_xticks(positions, YEARS)
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    ax.legend()
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "图1_年度利润比较.png", bbox_inches="tight")
    plt.close(fig)

    # 图2：各情形七年累计作物类别种植面积。
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), sharey=True)
    categories = ["豆类", "其他粮食", "其他蔬菜", "食用菌"]
    palette = ["#54A24B", "#ECA82C", "#4C78A8", "#B279A2"]
    for ax, result in zip(axes, results):
        df = result.crop_summary.copy()
        df["作物大类"] = df["作物类型"].map(crop_category)
        pivot = df.pivot_table(index="年份", columns="作物大类", values="种植面积/亩", aggfunc="sum", fill_value=0)
        bottom = np.zeros(len(YEARS))
        for category, color in zip(categories, palette):
            values = pivot.reindex(YEARS)[category].to_numpy() if category in pivot else np.zeros(len(YEARS))
            ax.bar(YEARS, values, bottom=bottom, label=category, color=color)
            bottom += values
        ax.set_title(result.scenario)
        ax.set_xlabel("年份")
        ax.grid(axis="y", linestyle="--", alpha=0.25)
    axes[0].set_ylabel("种植面积（亩·季）")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.suptitle("两种情形下的年度作物类别种植结构", y=0.98, fontsize=14)
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.94), ncol=4, frameon=False)
    fig.tight_layout(rect=(0, 0, 1, 0.88))
    fig.savefig(FIGURE_DIR / "图2_作物类别种植结构.png", bbox_inches="tight")
    plt.close(fig)

    # 图3：正常销量和超出销量。
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), sharey=True)
    for ax, result in zip(axes, results):
        annual = result.annual_summary
        normal = annual["正常销量/斤"].to_numpy() / 10000
        excess = annual["超出销量/斤"].to_numpy() / 10000
        ax.bar(YEARS, normal, label="正常销量", color="#4C78A8")
        ax.bar(YEARS, excess, bottom=normal, label="超出预期销售量部分", color="#F58518")
        ax.set_title(result.scenario)
        ax.set_xlabel("年份")
        ax.grid(axis="y", linestyle="--", alpha=0.25)
    axes[0].set_ylabel("产量（万斤）")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.suptitle("农作物正常销量与超出销量构成", y=0.98, fontsize=14)
    fig.legend(handles, labels, loc="upper center", bbox_to_anchor=(0.5, 0.94), ncol=2, frameon=False)
    fig.tight_layout(rect=(0, 0, 1, 0.88))
    fig.savefig(FIGURE_DIR / "图3_正常与超出销量.png", bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    np.random.seed(RANDOM_SEED)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    print("[STAGE A] 读取并校验附件数据……")
    data = load_input_data()
    print(f"  地块：{len(data.plots)} 个；作物：{len(data.crop_names)} 种；2023 记录：{len(data.planting_2023)} 条")

    scenarios = [
        ("超产部分滞销浪费", 0.0),
        ("超产部分按50%价格销售", 0.5),
    ]
    results = []
    for name, discount in scenarios:
        print(f"[STAGE B/C] 构建并求解：{name}……")
        result = build_and_solve(data, name, discount)
        results.append(result)
        print(f"  状态：{result.status}/{result.solution_status}；目标利润：{result.objective:,.2f} 元；耗时：{result.elapsed_seconds:.1f} 秒")

    print("[STAGE D] 写入竞赛模板和分析结果……")
    write_template(data, results[0], TEMPLATE_1, OUTPUT_DIR / "result1_1.xlsx")
    write_template(data, results[1], TEMPLATE_2, OUTPUT_DIR / "result1_2.xlsx")
    write_analysis_workbook(data, results, OUTPUT_DIR / "问题一_MILP分析结果.xlsx")
    write_text_log(results, LOG_DIR / "问题一_诊断日志.txt")

    summary_json = {
        result.scenario: {
            "状态": result.status,
            "整数解状态": result.solution_status,
            "目标利润_元": result.objective,
            "耗时_秒": result.elapsed_seconds,
        }
        for result in results
    }
    (LOG_DIR / "运行摘要.json").write_text(
        json.dumps(summary_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("[STAGE E] 绘制中文结果图……")
    create_figures(results)
    print("完成。结果位于 Q1/outputs、Q1/figures 和 Q1/logs。")


if __name__ == "__main__":
    main()
