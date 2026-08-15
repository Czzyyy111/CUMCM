"""2024 年高教社杯数学建模竞赛 C 题问题二：情景随机规划 + CVaR。

建模口径
--------
1. 2024—2030 年种植面积是所有情景共享的“事前决策”，避免完美信息泄漏；
2. 销量、亩产、成本和价格按题目趋势构造完整七年随机路径；
3. 训练集用拉丁超立方抽样，验证集和测试集用相互独立的蒙特卡洛抽样；
4. 主模型假设超产部分滞销、收入为 0；
5. 在期望利润与 90% 下尾 CVaR 之间权衡，并用独立验证集选取风险权重；
6. 最终方案在 5000 条全新测试情景上评价，所有原始附件只读。

默认运行会生成正式规模结果。可用 ``--quick`` 执行小规模流程检查；快速结果会在
输出文件和日志中明确标记，不应直接用于竞赛论文。
"""

from __future__ import annotations

import argparse
import json
import math
import shutil
import subprocess
import sys
import time
import warnings
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import pulp
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# =============================================================================
# 全局配置
# =============================================================================

YEARS = list(range(2024, 2031))
SEASONS = (1, 2)
BEAN_CROPS = set(range(1, 6)) | set(range(17, 20))
GRAIN_CROPS = set(range(1, 16))
VEGETABLE_CROPS = set(range(17, 38))
VEG_FIRST_CROPS = set(range(17, 35))
VEG_SECOND_D_CROPS = set(range(35, 38))
FUNGI_CROPS = set(range(38, 42))

MIN_AREA_OPEN = 5.0
MIN_AREA_GREENHOUSE = 0.3
MAX_PLOTS_PER_CROP = 7
CVaR_ALPHA = 0.90
MEAN_RETENTION = 0.97
CVaR_TIE_TOLERANCE = 0.0001
PREFERRED_RISK_WEIGHT = 0.4
LAMBDA_GRID = (0.0, 0.2, 0.4, 0.6, 0.8)
TRAIN_SCENARIOS = 200
OPTIMIZATION_SCENARIOS = 40
VALIDATION_SCENARIOS = 2000
TEST_SCENARIOS = 5000
TRAIN_SEED = 202402
VALIDATION_SEED = 202403
TEST_SEED = 202404
SOLVER_TIME_LIMIT = 75
SOLVER_GAP = 0.02
PROFIT_SCALE = 10_000.0  # 优化内部以万元计，兼顾 CVaR 边界与期望目标系数精度
EPS = 1e-6

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
DATA_DIR = ROOT_DIR / "C题"
OUTPUT_DIR = SCRIPT_DIR / "outputs"
LOG_DIR = SCRIPT_DIR / "logs"
SOURCE_DIR = SCRIPT_DIR / "source_data"
FIGURE_DIR = SCRIPT_DIR / "figures_nature"
ATTACHMENT_1 = DATA_DIR / "附件1.xlsx"
ATTACHMENT_2 = DATA_DIR / "附件2.xlsx"
TEMPLATE = DATA_DIR / "附件3" / "result2.xlsx"


@dataclass(frozen=True)
class Plot:
    name: str
    land_type: str
    area: float


@dataclass(frozen=True)
class Stat:
    crop_id: int
    crop_name: str
    land_type: str
    season_label: str
    yield_per_mu: float
    cost_per_mu: float
    price: float


@dataclass
class InputData:
    plots: dict[str, Plot]
    crop_names: dict[int, str]
    crop_types: dict[int, str]
    stats: dict[tuple[int, str, str], Stat]
    planting_2023: list[dict]
    sales_demand: dict[int, float]


@dataclass
class ScenarioSet:
    name: str
    method: str
    seed: int
    demand: np.ndarray       # [情景, 年份, 作物编号]
    yield_mult: np.ndarray   # [情景, 年份, 作物编号]
    cost_mult: np.ndarray
    price_mult: np.ndarray
    probabilities: np.ndarray | None = None

    @property
    def n(self) -> int:
        return int(self.demand.shape[0])

    @property
    def weights(self) -> np.ndarray:
        if self.probabilities is None:
            return np.full(self.n, 1.0 / self.n)
        weights = np.asarray(self.probabilities, dtype=float)
        if len(weights) != self.n or np.any(weights < 0) or not np.isclose(weights.sum(), 1.0):
            raise ValueError(f"情景 {self.name} 的概率权重无效。")
        return weights


@dataclass
class Candidate:
    risk_weight: float
    x_values: dict[tuple[str, int, int, int], float]
    y_values: dict[tuple[str, int, int, int], int]
    status: str
    solution_status: str
    objective: float
    elapsed_seconds: float
    mip_gap: float


@dataclass
class Evaluation:
    total_profit: np.ndarray
    annual_profit: np.ndarray
    total_revenue: np.ndarray
    total_cost: np.ndarray
    total_production: np.ndarray
    total_sales: np.ndarray
    unsold_rate: np.ndarray


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_price(value: object) -> float:
    text = clean_text(value).replace("—", "-").replace("–", "-")
    if "-" in text:
        low, high = text.split("-", maxsplit=1)
        return (float(low) + float(high)) / 2.0
    return float(text)


def season_label(raw: object) -> str:
    text = clean_text(raw)
    if "单季" in text:
        return "单季"
    if "第一" in text:
        return "第一季"
    if "第二" in text:
        return "第二季"
    raise ValueError(f"无法识别季次：{raw!r}")


# =============================================================================
# STAGE A 读取并校验输入
# =============================================================================

def load_input_data() -> InputData:
    for path in (ATTACHMENT_1, ATTACHMENT_2, TEMPLATE):
        if not path.exists():
            raise FileNotFoundError(f"缺少输入文件：{path}")

    wb1 = load_workbook(ATTACHMENT_1, data_only=True, read_only=True)
    plots = {
        clean_text(row[0]): Plot(clean_text(row[0]), clean_text(row[1]), float(row[2]))
        for row in wb1["乡村的现有耕地"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    crop_names: dict[int, str] = {}
    crop_types: dict[int, str] = {}
    for row in wb1["乡村种植的农作物"].iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], (int, float)):
            crop_names[int(row[0])] = clean_text(row[1])
            crop_types[int(row[0])] = clean_text(row[2])
    wb1.close()

    wb2 = load_workbook(ATTACHMENT_2, data_only=True, read_only=True)
    planting_2023: list[dict] = []
    current_plot = ""
    for row in wb2["2023年的农作物种植情况"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            current_plot = clean_text(row[0])
        if current_plot and isinstance(row[1], (int, float)):
            planting_2023.append({
                "plot": current_plot,
                "crop_id": int(row[1]),
                "crop_name": clean_text(row[2]),
                "area": float(row[4]),
                "season": season_label(row[5]),
            })

    stats: dict[tuple[int, str, str], Stat] = {}
    for row in wb2["2023年统计的相关数据"].iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        stat = Stat(
            crop_id=int(row[1]), crop_name=clean_text(row[2]),
            land_type=clean_text(row[3]), season_label=season_label(row[4]),
            yield_per_mu=float(row[5]), cost_per_mu=float(row[6]),
            price=parse_price(row[7]),
        )
        stats[(stat.crop_id, stat.land_type, stat.season_label)] = stat
    wb2.close()

    if len(plots) != 54 or len(crop_names) != 41 or len(planting_2023) != 87:
        raise ValueError(
            f"基础数据规模异常：地块={len(plots)}，作物={len(crop_names)}，"
            f"2023记录={len(planting_2023)}"
        )
    data = InputData(plots, crop_names, crop_types, stats, planting_2023, {})
    demand = defaultdict(float)
    for row in planting_2023:
        plot = plots[row["plot"]]
        season = 1 if row["season"] in {"单季", "第一季"} else 2
        demand[row["crop_id"]] += row["area"] * get_stat(data, plot, row["crop_id"], season).yield_per_mu
    data.sales_demand = {crop: float(demand.get(crop, 0.0)) for crop in crop_names}
    missing = [crop for crop, value in data.sales_demand.items() if value <= 0]
    if missing:
        raise ValueError(f"无法构造销量基准的作物：{missing}")
    return data


def get_stat(data: InputData, plot: Plot, crop_id: int, season: int) -> Stat:
    if plot.land_type in {"平旱地", "梯田", "山坡地"}:
        key = (crop_id, plot.land_type, "单季")
    elif plot.land_type == "水浇地":
        key = (crop_id, "水浇地", "单季" if crop_id == 16 else f"第{'一' if season == 1 else '二'}季")
    elif plot.land_type == "普通大棚":
        key = (crop_id, "普通大棚", f"第{'一' if season == 1 else '二'}季")
    elif plot.land_type == "智慧大棚":
        key = (crop_id, "普通大棚", "第一季") if season == 1 else (crop_id, "智慧大棚", "第二季")
    else:
        raise KeyError(f"未知地块类型：{plot.land_type}")
    if key not in data.stats:
        raise KeyError(f"缺少统计参数：{key}，地块={plot.name}，季次={season}")
    return data.stats[key]


def allowed_crops(plot: Plot, season: int) -> list[int]:
    if plot.land_type in {"平旱地", "梯田", "山坡地"}:
        return sorted(GRAIN_CROPS) if season == 1 else []
    if plot.land_type == "水浇地":
        return sorted({16} | VEG_FIRST_CROPS) if season == 1 else sorted(VEG_SECOND_D_CROPS)
    if plot.land_type == "普通大棚":
        return sorted(VEG_FIRST_CROPS if season == 1 else FUNGI_CROPS)
    if plot.land_type == "智慧大棚":
        return sorted(VEG_FIRST_CROPS)
    return []


def minimum_area(plot: Plot) -> float:
    return MIN_AREA_GREENHOUSE if "大棚" in plot.land_type else min(MIN_AREA_OPEN, plot.area)


def fixed_2023_bean_area(data: InputData, plot_name: str) -> float:
    return sum(
        row["area"] for row in data.planting_2023
        if row["plot"] == plot_name and row["crop_id"] in BEAN_CROPS
    )


# =============================================================================
# STAGE B 生成训练、验证和测试情景
# =============================================================================

def unit_samples(n: int, dimensions: int, seed: int, method: str) -> np.ndarray:
    """生成 [0,1] 样本；LHS 对每一维覆盖全部等概率分层。"""
    rng = np.random.default_rng(seed)
    if method == "蒙特卡洛":
        return rng.random((n, dimensions))
    if method != "拉丁超立方":
        raise ValueError(f"未知抽样方法：{method}")
    samples = np.empty((n, dimensions), dtype=float)
    for j in range(dimensions):
        samples[:, j] = (rng.permutation(n) + rng.random(n)) / n
    return samples


def triangular_inverse(u: np.ndarray, low: float, mode: float, high: float) -> np.ndarray:
    split = (mode - low) / (high - low)
    left = low + np.sqrt(u * (high - low) * (mode - low))
    right = high - np.sqrt((1.0 - u) * (high - low) * (high - mode))
    return np.where(u < split, left, right)


def generate_scenarios(data: InputData, n: int, seed: int, method: str, name: str) -> ScenarioSet:
    crops = sorted(data.crop_names)
    dimensions = len(YEARS) * len(crops) * 4
    u = unit_samples(n, dimensions, seed, method).reshape(n, len(YEARS), len(crops), 4)
    demand = np.zeros((n, len(YEARS), 42), dtype=float)
    yield_mult = np.ones_like(demand)
    cost_mult = np.ones_like(demand)
    price_mult = np.ones_like(demand)

    for c_idx, crop in enumerate(crops):
        current_demand = np.full(n, data.sales_demand[crop], dtype=float)
        current_cost = np.ones(n, dtype=float)
        current_price = np.ones(n, dtype=float)
        is_morel = "羊肚菌" in data.crop_names[crop]
        for t_idx, _year in enumerate(YEARS):
            u_d, u_y, u_c, u_p = (u[:, t_idx, c_idx, j] for j in range(4))
            if crop in {6, 7}:
                current_demand = current_demand * (1.0 + 0.05 + 0.05 * u_d)
                demand[:, t_idx, crop] = current_demand
            else:
                demand[:, t_idx, crop] = data.sales_demand[crop] * (0.95 + 0.10 * u_d)

            yield_mult[:, t_idx, crop] = 0.90 + 0.20 * u_y
            cost_growth = triangular_inverse(u_c, 0.03, 0.05, 0.07)
            current_cost = current_cost * (1.0 + cost_growth)
            cost_mult[:, t_idx, crop] = current_cost

            if crop <= 16:
                price_mult[:, t_idx, crop] = 1.0
            elif crop in VEGETABLE_CROPS:
                price_growth = triangular_inverse(u_p, 0.03, 0.05, 0.07)
                current_price = current_price * (1.0 + price_growth)
                price_mult[:, t_idx, crop] = current_price
            else:
                decline = np.full(n, 0.05) if is_morel else 0.01 + 0.04 * u_p
                current_price = current_price * (1.0 - decline)
                price_mult[:, t_idx, crop] = current_price

    return ScenarioSet(name, method, seed, demand, yield_mult, cost_mult, price_mult)


def reduce_training_scenarios(
    data: InputData, scenarios: ScenarioSet, target_n: int
) -> tuple[ScenarioSet, pd.DataFrame]:
    """按综合压力等频分层，并从每层选取最接近多维中心的真实路径。"""
    if target_n >= scenarios.n:
        mapping = pd.DataFrame({
            "原训练情景编号": np.arange(1, scenarios.n + 1),
            "代表情景编号": np.arange(1, scenarios.n + 1),
            "代表情景原编号": np.arange(1, scenarios.n + 1),
            "到代表情景标准化距离": np.zeros(scenarios.n),
            "代表情景概率": np.full(scenarios.n, 1.0 / scenarios.n),
        })
        return scenarios, mapping

    crops = np.asarray(sorted(data.crop_names), dtype=int)
    base_demand = np.asarray([data.sales_demand[int(c)] for c in crops], dtype=float)
    demand_ratio = scenarios.demand[:, :, crops] / base_demand[None, None, :]
    feature_blocks = [
        demand_ratio,
        scenarios.yield_mult[:, :, crops],
        scenarios.cost_mult[:, :, crops],
        scenarios.price_mult[:, :, crops],
    ]
    raw_features = np.concatenate(
        [block.reshape(scenarios.n, -1) for block in feature_blocks], axis=1
    )
    feature_std = raw_features.std(axis=0)
    usable = feature_std > 1e-10
    features = (raw_features[:, usable] - raw_features[:, usable].mean(axis=0)) / feature_std[usable]

    # 综合压力越低，需求、产量、价格越不利或成本越高。
    stress = (
        demand_ratio.mean(axis=(1, 2))
        + scenarios.yield_mult[:, :, crops].mean(axis=(1, 2))
        + scenarios.price_mult[:, :, crops].mean(axis=(1, 2))
        - scenarios.cost_mult[:, :, crops].mean(axis=(1, 2))
    )
    ordered = np.argsort(stress)
    strata = np.array_split(ordered, target_n)
    selected: list[int] = []
    cluster = np.empty(scenarios.n, dtype=int)
    representative_distance = np.empty(scenarios.n, dtype=float)
    probabilities = np.empty(target_n, dtype=float)
    for cluster_id, members in enumerate(strata):
        centroid = features[members].mean(axis=0)
        member_sq_dist = np.sum((features[members] - centroid) ** 2, axis=1)
        representative = int(members[np.argmin(member_sq_dist)])
        selected.append(representative)
        cluster[members] = cluster_id
        representative_distance[members] = np.sqrt(
            np.sum((features[members] - features[representative]) ** 2, axis=1)
        )
        probabilities[cluster_id] = len(members) / scenarios.n

    selected_array = np.asarray(selected, dtype=int)
    reduced = ScenarioSet(
        name=f"{scenarios.name}代表情景",
        method=f"{scenarios.method}+综合压力等频分层情景削减",
        seed=scenarios.seed,
        demand=scenarios.demand[selected_array].copy(),
        yield_mult=scenarios.yield_mult[selected_array].copy(),
        cost_mult=scenarios.cost_mult[selected_array].copy(),
        price_mult=scenarios.price_mult[selected_array].copy(),
        probabilities=probabilities,
    )
    mapping = pd.DataFrame({
        "原训练情景编号": np.arange(1, scenarios.n + 1),
        "代表情景编号": cluster + 1,
        "代表情景原编号": selected_array[cluster] + 1,
        "到代表情景标准化距离": representative_distance,
        "代表情景概率": probabilities[cluster],
    })
    return reduced, mapping


# =============================================================================
# STAGE C 构建情景随机 MILP，并求解风险权重候选
# =============================================================================

def build_stochastic_model(data: InputData, scenarios: ScenarioSet):
    model = pulp.LpProblem("Q2_Stochastic_CVaR", pulp.LpMaximize)
    x: dict[tuple[str, int, int, int], pulp.LpVariable] = {}
    y: dict[tuple[str, int, int, int], pulp.LpVariable] = {}
    stat_for_key: dict[tuple[str, int, int, int], Stat] = {}

    for plot in data.plots.values():
        for year in YEARS:
            for season in SEASONS:
                for crop in allowed_crops(plot, season):
                    key = (plot.name, year, season, crop)
                    x[key] = pulp.LpVariable(f"x_{plot.name}_{year}_{season}_{crop}", lowBound=0)
                    y[key] = pulp.LpVariable(f"y_{plot.name}_{year}_{season}_{crop}", cat="Binary")
                    stat_for_key[key] = get_stat(data, plot, crop, season)
                    model += x[key] <= plot.area * y[key], f"area_up_{plot.name}_{year}_{season}_{crop}"
                    model += x[key] >= minimum_area(plot) * y[key], f"area_lo_{plot.name}_{year}_{season}_{crop}"

    water_mode: dict[tuple[str, int], pulp.LpVariable] = {}
    for plot in data.plots.values():
        for year in YEARS:
            if plot.land_type == "水浇地":
                z = pulp.LpVariable(f"rice_mode_{plot.name}_{year}", cat="Binary")
                water_mode[(plot.name, year)] = z
                model += x[(plot.name, year, 1, 16)] == plot.area * z
                model += pulp.lpSum(x[(plot.name, year, 1, c)] for c in VEG_FIRST_CROPS) == plot.area * (1 - z)
                model += pulp.lpSum(x[(plot.name, year, 2, c)] for c in VEG_SECOND_D_CROPS) == plot.area * (1 - z)
            else:
                for season in SEASONS:
                    crops = allowed_crops(plot, season)
                    if crops:
                        model += pulp.lpSum(x[(plot.name, year, season, c)] for c in crops) == plot.area

    actual_2023 = defaultdict(set)
    for row in data.planting_2023:
        s_num = 1 if row["season"] in {"单季", "第一季"} else 2
        actual_2023[(row["plot"], s_num)].add(row["crop_id"])
    for plot in data.plots.values():
        if plot.land_type in {"平旱地", "梯田", "山坡地"}:
            for crop in GRAIN_CROPS:
                if crop in actual_2023[(plot.name, 1)]:
                    model += y[(plot.name, 2024, 1, crop)] == 0
                for year in YEARS[:-1]:
                    model += y[(plot.name, year, 1, crop)] + y[(plot.name, year + 1, 1, crop)] <= 1
        elif plot.land_type == "水浇地":
            if 16 in actual_2023[(plot.name, 1)]:
                model += y[(plot.name, 2024, 1, 16)] == 0
            for year in YEARS[:-1]:
                model += y[(plot.name, year, 1, 16)] + y[(plot.name, year + 1, 1, 16)] <= 1
        elif plot.land_type == "智慧大棚":
            for crop in VEG_FIRST_CROPS:
                if crop in actual_2023[(plot.name, 2)]:
                    model += y[(plot.name, 2024, 1, crop)] == 0
                for year in YEARS:
                    model += y[(plot.name, year, 1, crop)] + y[(plot.name, year, 2, crop)] <= 1
                for year in YEARS[:-1]:
                    model += y[(plot.name, year, 2, crop)] + y[(plot.name, year + 1, 1, crop)] <= 1

    for plot in data.plots.values():
        fixed = fixed_2023_bean_area(data, plot.name)
        for start in range(2023, 2029):
            terms = [
                x[key] for key in x
                if key[0] == plot.name and max(2024, start) <= key[1] < start + 3 and key[3] in BEAN_CROPS
            ]
            model += pulp.lpSum(terms) + (fixed if start == 2023 else 0.0) >= plot.area

    for year in YEARS:
        for season in SEASONS:
            for crop in data.crop_names:
                keys = [key for key in y if key[1] == year and key[2] == season and key[3] == crop]
                if keys:
                    model += pulp.lpSum(y[key] for key in keys) <= MAX_PLOTS_PER_CROP

    production_terms: dict[tuple[int, int, float], list] = defaultdict(list)
    for key, var in x.items():
        _, year, _, crop = key
        stat = stat_for_key[key]
        production_terms[(crop, year, stat.price)].append(stat.yield_per_mu * var)
    production_expr = {group: pulp.lpSum(terms) for group, terms in production_terms.items()}

    profit_expr: list[pulp.LpAffineExpression] = []
    groups_by_crop_year: dict[tuple[int, int], list[tuple[int, int, float]]] = defaultdict(list)
    for group in production_expr:
        groups_by_crop_year[(group[0], group[1])].append(group)

    for s in range(scenarios.n):
        revenue_terms = []
        sales_by_group: dict[tuple[int, int, float], pulp.LpVariable] = {}
        for group, expr in production_expr.items():
            crop, year, price = group
            t = year - YEARS[0]
            sale = pulp.LpVariable(f"q_{s}_{crop}_{year}_{price:.4f}", lowBound=0)
            sales_by_group[group] = sale
            model += sale <= scenarios.yield_mult[s, t, crop] * expr
            revenue_terms.append(price * scenarios.price_mult[s, t, crop] * sale / PROFIT_SCALE)
        for (crop, year), groups in groups_by_crop_year.items():
            t = year - YEARS[0]
            sales = [sales_by_group[g] for g in groups]
            model += pulp.lpSum(sales) <= scenarios.demand[s, t, crop]
        cost_terms = []
        for key, var in x.items():
            _, year, _, crop = key
            t = year - YEARS[0]
            cost_terms.append(
                stat_for_key[key].cost_per_mu * scenarios.cost_mult[s, t, crop] * var / PROFIT_SCALE
            )
        profit_expr.append(pulp.lpSum(revenue_terms) - pulp.lpSum(cost_terms))

    # 单位为万元；±20000 足以覆盖本题七年利润，同时避免原始“元”尺度的大边界。
    eta = pulp.LpVariable("cvar_eta", lowBound=-20_000.0, upBound=20_000.0)
    shortfalls = []
    for s, profit in enumerate(profit_expr):
        xi = pulp.LpVariable(f"cvar_shortfall_{s}", lowBound=0)
        model += xi >= eta - profit
        shortfalls.append(xi)
    scenario_weights = scenarios.weights
    expected_profit = pulp.lpSum(
        scenario_weights[s] * profit_expr[s] for s in range(scenarios.n)
    )
    lower_cvar = eta - pulp.lpSum(
        scenario_weights[s] * shortfalls[s] for s in range(scenarios.n)
    ) / (1.0 - CVaR_ALPHA)
    return model, x, y, stat_for_key, expected_profit, lower_cvar, len(water_mode)


def solve_with_optional_mip_start(
    model: pulp.LpProblem,
    solver: pulp.HiGHS,
    x: dict[tuple[str, int, int, int], pulp.LpVariable],
    y: dict[tuple[str, int, int, int], pulp.LpVariable],
    mip_start: Candidate | None,
) -> str:
    """调用 HiGHS；可提供第一阶段可行方案作为启动点，但不固定任何变量。"""
    if mip_start is None:
        model.solve(solver)
        return "无（独立冷启动）"

    # PuLP 的 HiGHS 接口尚未自动传递初始值，因此按其 actualSolve 流程构建模型，
    # 再将 x/y 作为部分 MIP start 交给 HiGHS。求解阶段仍保留全部变量的原始上下界。
    solver.createAndConfigureSolver(model)
    solver.buildSolverModel(model)
    start_variables = [*x.values(), *y.values()]
    start_indices = np.asarray([var.index for var in start_variables], dtype=np.int32)
    start_values = np.asarray(
        [
            mip_start.x_values[key] for key in x
        ] + [
            mip_start.y_values[key] for key in y
        ],
        dtype=np.float64,
    )
    start_status = model.solverModel.setSolution(
        len(start_indices), start_indices, start_values
    )
    solver.callSolver(model)
    status, solution_status = solver.findSolutionValues(model)
    for var in model.variables():
        var.modified = False
    for constraint in model.constraints.values():
        constraint.modifier = False
    model.assignStatus(status, solution_status)
    return f"λ={mip_start.risk_weight:.1f}第一阶段方案（{start_status}）"


def solve_candidates(data: InputData, scenarios: ScenarioSet, lambdas: tuple[float, ...], time_limit: int):
    candidates: list[Candidate] = []
    diagnostics: list[dict] = []
    common_stat_for_key: dict[tuple[str, int, int, int], Stat] | None = None
    risk_neutral_start: Candidate | None = None
    for candidate_index, risk_weight in enumerate(lambdas):
        # 每个风险权重必须拥有独立、完整的第一阶段决策空间。不能把 λ=0 的
        # 作物启用变量固定后再求其他权重，否则面积等式会使所有候选退化为同一方案。
        model, x, y, stat_for_key, expected_profit, lower_cvar, water_binary_count = (
            build_stochastic_model(data, scenarios)
        )
        if common_stat_for_key is None:
            common_stat_for_key = stat_for_key
        elif common_stat_for_key.keys() != stat_for_key.keys():
            raise RuntimeError("不同风险权重的决策变量集合不一致。")
        model.setObjective((1.0 - risk_weight) * expected_profit + risk_weight * lower_cvar)
        solver_log = LOG_DIR / f"HiGHS_风险权重_{risk_weight:.1f}.log"
        solver = pulp.HiGHS(
            msg=True, timeLimit=time_limit, gapRel=SOLVER_GAP, threads=4,
            log_file=str(solver_log),
        )
        if not solver.available():
            raise RuntimeError("PuLP 未找到 HiGHS，请安装 requirements.txt 中的 highspy。")
        start = time.perf_counter()
        mip_start_note = solve_with_optional_mip_start(
            model, solver, x, y, risk_neutral_start
        )
        elapsed = time.perf_counter() - start
        pulp_status = pulp.LpStatus.get(model.status, str(model.status))
        solution_status = pulp.LpSolution.get(model.sol_status, str(model.sol_status))
        highs_status = model.solverModel.modelStatusToString(model.solverModel.getModelStatus())
        info = model.solverModel.getInfo()
        if model.sol_status not in {pulp.LpSolutionOptimal, pulp.LpSolutionIntegerFeasible}:
            raise RuntimeError(
                f"风险权重 {risk_weight:.1f} 未得到整数可行解：{highs_status}/{solution_status}"
            )
        status = "达到时间上限，已有整数可行解" if "Time limit" in highs_status else highs_status
        candidate = Candidate(
            risk_weight=risk_weight,
            x_values={key: max(0.0, float(pulp.value(var) or 0.0)) for key, var in x.items()},
            y_values={key: int(round(float(pulp.value(var) or 0.0))) for key, var in y.items()},
            status=status, solution_status=solution_status,
            objective=float(pulp.value(model.objective)) * PROFIT_SCALE, elapsed_seconds=elapsed,
            mip_gap=float(info.mip_gap),
        )
        candidates.append(candidate)
        diagnostics.append({
            "风险权重": risk_weight, "PuLP状态": pulp_status, "整数解状态": solution_status,
            "HiGHS状态": highs_status, "目标函数值/元": candidate.objective,
            "最终相对间隙": candidate.mip_gap, "求解耗时/秒": elapsed,
            "变量总数": len(model.variables()), "约束总数": len(model.constraints),
            "二进制变量数": len(y) + water_binary_count,
            "支持集策略": "独立完整MILP；未固定其他风险权重的作物支持集",
            "MIP启动点": mip_start_note,
        })
        print(
            f"  λ={risk_weight:.1f}：{status}，目标={candidate.objective:,.0f} 元，"
            f"间隙={candidate.mip_gap:.2%}，耗时={elapsed:.1f} 秒"
        )
        if candidate_index == 0:
            risk_neutral_start = candidate
    if common_stat_for_key is None:
        raise RuntimeError("未生成任何风险权重候选。")
    return candidates, common_stat_for_key, pd.DataFrame(diagnostics)


# =============================================================================
# STAGE D 固定方案后的独立验证与测试
# =============================================================================

def plan_components(candidate: Candidate, stat_for_key: dict):
    production: dict[tuple[int, int, float], float] = defaultdict(float)
    costs: dict[tuple[int, int], float] = defaultdict(float)
    for key, area in candidate.x_values.items():
        if area <= EPS:
            continue
        _, year, _, crop = key
        stat = stat_for_key[key]
        production[(crop, year, stat.price)] += area * stat.yield_per_mu
        costs[(crop, year)] += area * stat.cost_per_mu
    return production, costs


def evaluate_plan(candidate: Candidate, stat_for_key: dict, scenarios: ScenarioSet) -> Evaluation:
    production, costs = plan_components(candidate, stat_for_key)
    annual_revenue = np.zeros((scenarios.n, len(YEARS)))
    annual_cost = np.zeros_like(annual_revenue)
    annual_production = np.zeros_like(annual_revenue)
    annual_sales = np.zeros_like(annual_revenue)
    groups = defaultdict(list)
    for (crop, year, price), base_production in production.items():
        groups[(crop, year)].append((price, base_production))

    for (crop, year), crop_groups in groups.items():
        t = year - YEARS[0]
        remaining = scenarios.demand[:, t, crop].copy()
        for price, base_production in sorted(crop_groups, reverse=True):
            realized = base_production * scenarios.yield_mult[:, t, crop]
            sold = np.minimum(realized, remaining)
            annual_revenue[:, t] += sold * price * scenarios.price_mult[:, t, crop]
            annual_production[:, t] += realized
            annual_sales[:, t] += sold
            remaining -= sold
        annual_cost[:, t] += costs[(crop, year)] * scenarios.cost_mult[:, t, crop]

    annual_profit = annual_revenue - annual_cost
    total_production = annual_production.sum(axis=1)
    total_sales = annual_sales.sum(axis=1)
    unsold_rate = np.divide(
        total_production - total_sales, total_production,
        out=np.zeros_like(total_production), where=total_production > 0,
    )
    return Evaluation(
        total_profit=annual_profit.sum(axis=1), annual_profit=annual_profit,
        total_revenue=annual_revenue.sum(axis=1), total_cost=annual_cost.sum(axis=1),
        total_production=total_production, total_sales=total_sales,
        unsold_rate=unsold_rate,
    )


def lower_tail_cvar(values: np.ndarray, alpha: float = CVaR_ALPHA) -> float:
    count = max(1, int(math.ceil((1.0 - alpha) * len(values))))
    return float(np.sort(values)[:count].mean())


def evaluation_metrics(evaluation: Evaluation, prefix: str = "") -> dict:
    p = evaluation.total_profit
    return {
        f"{prefix}情景数": len(p),
        f"{prefix}期望利润/元": float(p.mean()),
        f"{prefix}利润标准差/元": float(p.std(ddof=1)),
        f"{prefix}P10利润/元": float(np.quantile(p, 0.10)),
        f"{prefix}P50利润/元": float(np.quantile(p, 0.50)),
        f"{prefix}P90利润/元": float(np.quantile(p, 0.90)),
        f"{prefix}下尾CVaR90/元": lower_tail_cvar(p),
        f"{prefix}累计亏损概率": float(np.mean(p < 0)),
        f"{prefix}平均滞销率": float(evaluation.unsold_rate.mean()),
    }


def training_objective_table(
    candidates: list[Candidate], train_evals: dict[float, Evaluation]
) -> pd.DataFrame:
    """按论文公式在实际进入优化的训练情景上重算目标。"""
    rows = []
    for candidate in candidates:
        evaluation = train_evals[candidate.risk_weight]
        mean_profit = float(evaluation.total_profit.mean())
        cvar_profit = lower_tail_cvar(evaluation.total_profit)
        weighted_objective = (
            (1.0 - candidate.risk_weight) * mean_profit
            + candidate.risk_weight * cvar_profit
        )
        rows.append({
            "风险权重λ": candidate.risk_weight,
            "训练集_期望利润/元": mean_profit,
            "训练集_下尾CVaR90/元": cvar_profit,
            "训练集_CVaR加权目标/元": weighted_objective,
            "求解器目标函数值/元": candidate.objective,
            "目标重算差异/元": weighted_objective - candidate.objective,
        })
    return pd.DataFrame(rows).sort_values("风险权重λ").reset_index(drop=True)


def plan_difference_table(candidates: list[Candidate]) -> pd.DataFrame:
    """比较候选方案，防止风险权重变化但种植决策被意外固定。"""
    baseline = min(candidates, key=lambda c: c.risk_weight)
    rows = []
    for candidate in sorted(candidates, key=lambda c: c.risk_weight):
        area_diffs = np.array([
            abs(candidate.x_values[key] - baseline.x_values[key])
            for key in baseline.x_values
        ])
        support_changes = sum(
            candidate.y_values[key] != baseline.y_values[key]
            for key in baseline.y_values
        )
        rows.append({
            "风险权重λ": candidate.risk_weight,
            "比较基准λ": baseline.risk_weight,
            "面积变化单元数": int(np.sum(area_diffs > 1e-4)),
            "作物支持集变化单元数": int(support_changes),
            "总绝对面积差/亩": float(area_diffs.sum()),
            "最大面积差/亩": float(area_diffs.max(initial=0.0)),
            "是否与基准方案相同": "是" if np.all(area_diffs <= 1e-4) and support_changes == 0 else "否",
        })
    return pd.DataFrame(rows)


def choose_candidate(candidates: list[Candidate], validation_evals: dict[float, Evaluation]):
    rows = []
    for candidate in candidates:
        metrics = evaluation_metrics(validation_evals[candidate.risk_weight], "验证集_")
        rows.append({
            "风险权重λ": candidate.risk_weight,
            "求解状态": candidate.status,
            "MIP相对间隙": candidate.mip_gap,
            "求解耗时/秒": candidate.elapsed_seconds,
            **metrics,
        })
    frontier = pd.DataFrame(rows).sort_values("风险权重λ").reset_index(drop=True)
    best_mean = float(frontier["验证集_期望利润/元"].max())
    eligible = frontier[frontier["验证集_期望利润/元"] >= MEAN_RETENTION * best_mean].copy()
    best_cvar = float(eligible["验证集_下尾CVaR90/元"].max())
    near_best = eligible[
        eligible["验证集_下尾CVaR90/元"] >= best_cvar - CVaR_TIE_TOLERANCE * max(abs(best_cvar), 1.0)
    ].copy()
    near_best["距预设中等风险权重"] = (near_best["风险权重λ"] - PREFERRED_RISK_WEIGHT).abs()
    chosen_row = near_best.sort_values(["距预设中等风险权重", "风险权重λ"]).iloc[0]
    chosen_lambda = float(chosen_row["风险权重λ"])
    selected = next(c for c in candidates if math.isclose(c.risk_weight, chosen_lambda))
    frontier["是否入选"] = np.where(np.isclose(frontier["风险权重λ"], chosen_lambda), "是", "否")
    frontier["选择规则"] = (
        f"验证集期望利润不低于最高值的{MEAN_RETENTION:.0%}；CVaR在{CVaR_TIE_TOLERANCE:.2%}"
        f"容差内并列时优先中等风险权重{PREFERRED_RISK_WEIGHT:.1f}"
    )
    return selected, frontier


def validate_hard_constraints(data: InputData, candidate: Candidate) -> dict:
    x, y = candidate.x_values, candidate.y_values
    max_area_error = 0.0
    min_area_violations = dispersion_violations = bean_violations = rotation_violations = 0
    for key, value in x.items():
        plot = data.plots[key[0]]
        if value > EPS and value + 1e-5 < minimum_area(plot):
            min_area_violations += 1
        if y[key] == 0 and value > 1e-5:
            min_area_violations += 1
    for plot in data.plots.values():
        for year in YEARS:
            if plot.land_type == "水浇地":
                rice = x[(plot.name, year, 1, 16)]
                veg1 = sum(x[(plot.name, year, 1, c)] for c in VEG_FIRST_CROPS)
                veg2 = sum(x[(plot.name, year, 2, c)] for c in VEG_SECOND_D_CROPS)
                if rice > plot.area / 2:
                    max_area_error = max(max_area_error, abs(rice - plot.area), veg1, veg2)
                else:
                    max_area_error = max(max_area_error, rice, abs(veg1 - plot.area), abs(veg2 - plot.area))
            else:
                for season in SEASONS:
                    crops = allowed_crops(plot, season)
                    if crops:
                        max_area_error = max(
                            max_area_error,
                            abs(sum(x[(plot.name, year, season, c)] for c in crops) - plot.area),
                        )
        fixed = fixed_2023_bean_area(data, plot.name)
        for start in range(2023, 2029):
            bean_area = fixed if start == 2023 else 0.0
            bean_area += sum(
                value for key, value in x.items()
                if key[0] == plot.name and max(2024, start) <= key[1] < start + 3 and key[3] in BEAN_CROPS
            )
            bean_violations += int(bean_area + 1e-5 < plot.area)
    for year in YEARS:
        for season in SEASONS:
            for crop in data.crop_names:
                count = sum(v for key, v in y.items() if key[1:] == (year, season, crop))
                dispersion_violations += int(count > MAX_PLOTS_PER_CROP)
    for plot in data.plots.values():
        if plot.land_type in {"平旱地", "梯田", "山坡地"}:
            for crop in GRAIN_CROPS:
                for year in YEARS[:-1]:
                    rotation_violations += int(y[(plot.name, year, 1, crop)] + y[(plot.name, year + 1, 1, crop)] > 1)
        elif plot.land_type == "水浇地":
            for year in YEARS[:-1]:
                rotation_violations += int(y[(plot.name, year, 1, 16)] + y[(plot.name, year + 1, 1, 16)] > 1)
        elif plot.land_type == "智慧大棚":
            for crop in VEG_FIRST_CROPS:
                for year in YEARS:
                    rotation_violations += int(y[(plot.name, year, 1, crop)] + y[(plot.name, year, 2, crop)] > 1)
                for year in YEARS[:-1]:
                    rotation_violations += int(y[(plot.name, year, 2, crop)] + y[(plot.name, year + 1, 1, crop)] > 1)
    return {
        "面积平衡最大误差/亩": max_area_error,
        "最小面积约束违规数": min_area_violations,
        "分散度约束违规数": dispersion_violations,
        "豆类轮作约束违规数": bean_violations,
        "重茬约束违规数": rotation_violations,
    }


# =============================================================================
# STAGE E 输出模板、结果表、源数据和诊断日志
# =============================================================================

def write_template(data: InputData, candidate: Candidate, output: Path) -> None:
    shutil.copy2(TEMPLATE, output)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(output)
    for year in YEARS:
        ws = wb[str(year)]
        crop_col = {
            clean_text(ws.cell(1, col).value): col for col in range(3, ws.max_column + 1)
            if ws.cell(1, col).value is not None
        }
        for row_start, row_end, season in ((2, 55, 1), (56, 83, 2)):
            for row in range(row_start, row_end + 1):
                plot_name = clean_text(ws.cell(row, 2).value)
                if not plot_name:
                    continue
                for crop, crop_name in data.crop_names.items():
                    value = candidate.x_values.get((plot_name, year, season, crop), 0.0)
                    ws.cell(row, crop_col[crop_name]).value = 0 if value < 5e-5 else round(value, 4)
    wb.save(output)


def crop_category(crop_type: str) -> str:
    text = str(crop_type)
    if "豆类" in text:
        return "豆类"
    if "食用菌" in text:
        return "食用菌"
    if "蔬菜" in text:
        return "其他蔬菜"
    return "其他粮食"


def plan_tables(data: InputData, candidate: Candidate):
    details = []
    for key, area in candidate.x_values.items():
        if area <= 5e-5:
            continue
        plot, year, season, crop = key
        details.append({
            "年份": year, "季次": f"第{season}季", "地块名称": plot,
            "地块类型": data.plots[plot].land_type, "作物编号": crop,
            "作物名称": data.crop_names[crop], "作物类型": data.crop_types[crop],
            "种植面积/亩": area,
        })
    detail_df = pd.DataFrame(details).sort_values(["年份", "季次", "地块名称", "作物编号"])
    crop_df = detail_df.groupby(
        ["年份", "作物编号", "作物名称", "作物类型"], as_index=False
    )["种植面积/亩"].sum()
    crop_df["作物大类"] = crop_df["作物类型"].map(crop_category)
    structure = crop_df.groupby(["年份", "作物大类"], as_index=False)["种植面积/亩"].sum()
    return detail_df, crop_df, structure


def typical_scenarios(evaluation: Evaluation):
    labels = [("最差情景", 0.0), ("下行典型情景（P10）", 0.10), ("中位情景（P50）", 0.50),
              ("上行情景（P90）", 0.90), ("最佳情景", 1.0)]
    rows = []
    for label, quantile in labels:
        target = np.quantile(evaluation.total_profit, quantile)
        idx = int(np.argmin(np.abs(evaluation.total_profit - target)))
        for t, year in enumerate(YEARS):
            rows.append({
                "典型情景": label, "测试情景编号": idx + 1, "分位位置": quantile,
                "年份": year, "年度利润/元": evaluation.annual_profit[idx, t],
                "七年累计利润/元": evaluation.total_profit[idx],
            })
    return pd.DataFrame(rows)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2, 32)
            ws.column_dimensions[col[0].column_letter].width = max(width, 10)
    wb.save(path)


def available_output_path(preferred: Path, always_new_if_exists: bool = False) -> Path:
    """为可能被 Excel 占用的工作簿选择一个不会覆盖现有文件的新路径。"""
    if not preferred.exists() or not always_new_if_exists:
        return preferred
    candidate = preferred.with_name(f"{preferred.stem}_更新版{preferred.suffix}")
    serial = 2
    while candidate.exists():
        candidate = preferred.with_name(f"{preferred.stem}_更新版_{serial}{preferred.suffix}")
        serial += 1
    warnings.warn(
        f"为避免覆盖可能正在打开的 {preferred.name}，结果写至 {candidate.name}。",
        RuntimeWarning,
    )
    return candidate


def write_outputs(
    data: InputData, candidates: list[Candidate], selected: Candidate,
    frontier: pd.DataFrame, solver_diag: pd.DataFrame, test_evals: dict[float, Evaluation],
    selected_test: Evaluation, hard_checks: dict, quick: bool,
    training_objectives: pd.DataFrame, plan_differences: pd.DataFrame,
    scenario_reduction: pd.DataFrame,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    template_output = available_output_path(OUTPUT_DIR / "result2.xlsx")
    write_template(data, selected, template_output)

    detail_df, crop_df, structure_df = plan_tables(data, selected)
    typical_df = typical_scenarios(selected_test)
    totals_df = pd.DataFrame({
        "测试情景编号": np.arange(1, len(selected_test.total_profit) + 1),
        "七年累计利润/元": selected_test.total_profit,
        "七年销售收入/元": selected_test.total_revenue,
        "七年种植成本/元": selected_test.total_cost,
        "七年总产量/斤": selected_test.total_production,
        "七年正常销量/斤": selected_test.total_sales,
        "七年滞销率": selected_test.unsold_rate,
    })
    annual_wide = pd.DataFrame(selected_test.annual_profit, columns=[f"{year}年利润/元" for year in YEARS])
    annual_wide.insert(0, "测试情景编号", np.arange(1, len(annual_wide) + 1))

    test_frontier_rows = []
    for candidate in candidates:
        metrics = evaluation_metrics(test_evals[candidate.risk_weight], "测试集_")
        test_frontier_rows.append({"风险权重λ": candidate.risk_weight, **metrics})
    test_frontier = pd.DataFrame(test_frontier_rows)
    full_frontier = (
        frontier
        .merge(training_objectives, on="风险权重λ", how="left")
        .merge(test_frontier, on="风险权重λ", how="left")
    )

    annual_summary = []
    for t, year in enumerate(YEARS):
        values = selected_test.annual_profit[:, t]
        annual_summary.append({
            "年份": year, "平均利润/元": values.mean(), "利润标准差/元": values.std(ddof=1),
            "P10利润/元": np.quantile(values, 0.10), "P50利润/元": np.quantile(values, 0.50),
            "P90利润/元": np.quantile(values, 0.90), "年度亏损概率": np.mean(values < 0),
        })
    annual_summary_df = pd.DataFrame(annual_summary)

    assumptions = pd.DataFrame([
        ("小麦、玉米销量", "逐年增长率 U(5%,10%)，在同一情景路径中复合"),
        ("其他作物销量", "各年相对 2023 年基准 U(-5%,5%)，不复合"),
        ("亩产量", "各作物各年乘数 U(0.90,1.10)"),
        ("种植成本", "年增长率 Tri(3%,5%,7%)，逐年复合"),
        ("粮食价格", "保持 2023 年价格区间中点"),
        ("蔬菜价格", "年增长率 Tri(3%,5%,7%)，逐年复合"),
        ("食用菌价格", "除羊肚菌外年降幅 U(1%,5%)；羊肚菌固定下降 5%"),
        ("问题二独立性", "跨作物、跨参数独立抽样；相关性留待问题三"),
        ("超产处理", "主模型超产部分滞销，收入为 0"),
    ], columns=["随机因素", "分布及口径"])
    methods = pd.DataFrame([
        ("模型", "共享种植决策的情景随机混合整数线性规划 + 90%下尾CVaR"),
        ("计算策略", "每个风险权重分别构建并求解完整随机MILP；作物支持集和种植面积均重新优化"),
        ("训练池", f"{TRAIN_SCENARIOS if not quick else 20}条拉丁超立方七年路径；用于情景削减与内部重评价"),
        ("优化训练集", f"正式求解使用{OPTIMIZATION_SCENARIOS if not quick else 20}条综合压力等频分层代表路径，代表概率取每层训练路径频率；论文目标函数据此计算"),
        ("验证集", f"{VALIDATION_SCENARIOS if not quick else 200}条独立蒙特卡洛路径，用于选择风险权重"),
        ("测试集", f"{TEST_SCENARIOS if not quick else 500}条全新蒙特卡洛路径，仅用于最终评价"),
        ("非预见性", "所有训练情景共享同一套种植面积 x；测试情景中不重新优化"),
        ("选择规则", f"验证集期望利润至少保留{MEAN_RETENTION:.0%}，再最大化下尾CVaR90；"
                    f"CVaR在{CVaR_TIE_TOLERANCE:.2%}容差内并列时优先λ={PREFERRED_RISK_WEIGHT:.1f}"),
        ("最终风险权重", f"λ={selected.risk_weight:.1f}"),
        ("利润函数", "正常销售收入－种植成本；销量不超过随机产量与随机需求"),
        ("数据泄漏检查", "训练、验证、测试随机种子相互独立；未使用未来真实观测；测试中冻结方案"),
        ("运行标记", "快速流程检查（不可直接用于论文）" if quick else "正式规模运行"),
    ], columns=["项目", "说明"])
    diag_rows = [{"检查项": key, "结果": value} for key, value in hard_checks.items()]
    diag_rows.extend([
        {"检查项": "原始地块数", "结果": len(data.plots)},
        {"检查项": "原始作物数", "结果": len(data.crop_names)},
        {"检查项": "2023种植记录数", "结果": len(data.planting_2023)},
        {"检查项": "训练/验证/测试随机种子", "结果": f"{TRAIN_SEED}/{VALIDATION_SEED}/{TEST_SEED}"},
        {"检查项": "最终方案风险权重", "结果": selected.risk_weight},
        {"检查项": "正式结果标记", "结果": not quick},
    ])
    diag_df = pd.DataFrame(diag_rows)

    workbook = available_output_path(
        OUTPUT_DIR / "问题二_随机规划_CVaR分析结果.xlsx",
        always_new_if_exists=True,
    )
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        full_frontier.to_excel(writer, sheet_name="风险收益前沿", index=False)
        annual_summary_df.to_excel(writer, sheet_name="最终方案年度评估", index=False)
        totals_df.to_excel(writer, sheet_name="测试情景利润", index=False)
        annual_wide.to_excel(writer, sheet_name="测试情景年度利润", index=False)
        typical_df.to_excel(writer, sheet_name="典型情景", index=False)
        detail_df.to_excel(writer, sheet_name="地块种植方案", index=False)
        crop_df.to_excel(writer, sheet_name="作物面积汇总", index=False)
        structure_df.to_excel(writer, sheet_name="作物结构", index=False)
        assumptions.to_excel(writer, sheet_name="情景分布假设", index=False)
        methods.to_excel(writer, sheet_name="方法说明", index=False)
        solver_diag.to_excel(writer, sheet_name="求解诊断", index=False)
        training_objectives.to_excel(writer, sheet_name="训练目标函数", index=False)
        plan_differences.to_excel(writer, sheet_name="方案差异诊断", index=False)
        scenario_reduction.to_excel(writer, sheet_name="情景削减映射", index=False)
        diag_df.to_excel(writer, sheet_name="约束与泄漏检查", index=False)
    style_workbook(workbook)

    full_frontier.to_csv(SOURCE_DIR / "风险收益前沿.csv", index=False, encoding="utf-8-sig")
    training_objectives.to_csv(SOURCE_DIR / "训练目标函数.csv", index=False, encoding="utf-8-sig")
    plan_differences.to_csv(SOURCE_DIR / "方案差异诊断.csv", index=False, encoding="utf-8-sig")
    scenario_reduction.to_csv(SOURCE_DIR / "情景削减映射.csv", index=False, encoding="utf-8-sig")
    totals_df.to_csv(SOURCE_DIR / "测试情景利润.csv", index=False, encoding="utf-8-sig")
    annual_wide.to_csv(SOURCE_DIR / "测试情景年度利润.csv", index=False, encoding="utf-8-sig")
    typical_df.to_csv(SOURCE_DIR / "典型情景年度利润.csv", index=False, encoding="utf-8-sig")
    structure_df.to_csv(SOURCE_DIR / "年度作物结构.csv", index=False, encoding="utf-8-sig")

    summary = {
        "运行模式": "快速流程检查" if quick else "正式规模",
        "最终风险权重": selected.risk_weight,
        "优化代表情景数": int(scenario_reduction["代表情景编号"].nunique()),
        "分析工作簿": workbook.name,
        "官方结果文件": template_output.name,
        "候选方案是否全部重合": bool(plan_differences["是否与基准方案相同"].eq("是").all()),
        **evaluation_metrics(selected_test, "测试集_"),
        **hard_checks,
    }
    (LOG_DIR / "运行摘要.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = ["问题二 情景随机规划 + CVaR 诊断日志", "=" * 64, ""]
    for key, value in summary.items():
        lines.append(f"{key}: {value}")
    lines.extend([
        "", "数据隔离与有效性检查：",
        "1. 训练集使用拉丁超立方抽样，验证集和测试集使用不同种子的蒙特卡洛抽样。",
        "2. 风险权重只由验证集选择，5000 条正式测试情景不参与优化与选择。",
        "3. 所有情景共享种植决策；样本外评价时固定方案，不逐情景重新优化。",
        "4. 测试情景利润分布使用全部记录，绘图未抽样或删除观测。",
        "5. 分布参数来源于题面区间；成本和蔬菜价格的三角分布边界属于可审计工程假设。",
        "6. 每个CVaR风险权重独立求解完整随机MILP，作物支持集和种植面积均未跨权重固定。",
        "7. 方案差异表自动比较各权重与λ=0方案的面积和作物支持集，防止候选被意外锁定。",
        "8. 200条LHS训练池经综合压力等频分层压缩为40条等权代表情景求解；目标函数在40条优化训练情景上重算，完整200条仅作内部重评价。",
    ])
    (LOG_DIR / "问题二_诊断日志.txt").write_text("\n".join(lines), encoding="utf-8")


def parse_args():
    parser = argparse.ArgumentParser(description="问题二：情景随机规划 + CVaR")
    parser.add_argument("--quick", action="store_true", help="小规模流程检查，不作为正式论文结果")
    parser.add_argument("--skip-figures", action="store_true", help="跳过 Nature 风格图形生成")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for directory in (OUTPUT_DIR, LOG_DIR, SOURCE_DIR, FIGURE_DIR):
        directory.mkdir(parents=True, exist_ok=True)
    train_n, val_n, test_n = (
        (20, 200, 500) if args.quick else (TRAIN_SCENARIOS, VALIDATION_SCENARIOS, TEST_SCENARIOS)
    )
    lambdas = (0.0, 0.4, 0.8) if args.quick else LAMBDA_GRID
    time_limit = 35 if args.quick else SOLVER_TIME_LIMIT
    print("[STAGE A] 读取并校验附件数据……")
    data = load_input_data()
    print(f"  地块={len(data.plots)}，作物={len(data.crop_names)}，2023记录={len(data.planting_2023)}")

    print("[STAGE B] 构造相互独立的训练、验证和测试情景……")
    train = generate_scenarios(data, train_n, TRAIN_SEED, "拉丁超立方", "训练集")
    optimization_train, scenario_reduction = reduce_training_scenarios(
        data, train, min(OPTIMIZATION_SCENARIOS, train.n)
    )
    validation = generate_scenarios(data, val_n, VALIDATION_SEED, "蒙特卡洛", "验证集")
    test = generate_scenarios(data, test_n, TEST_SEED, "蒙特卡洛", "测试集")
    print(
        f"  训练={train.n}（优化代表情景={optimization_train.n}），"
        f"验证={validation.n}，测试={test.n}"
    )

    print("[STAGE C] 构建随机 MILP 并求解风险权重候选……")
    candidates, stat_for_key, solver_diag = solve_candidates(
        data, optimization_train, lambdas, time_limit
    )

    print("[STAGE D] 在独立验证集选择风险权重，并在全新测试集评价……")
    optimization_evals = {
        c.risk_weight: evaluate_plan(c, stat_for_key, optimization_train) for c in candidates
    }
    full_training_evals = {
        c.risk_weight: evaluate_plan(c, stat_for_key, train) for c in candidates
    }
    training_objectives = training_objective_table(candidates, optimization_evals)
    full_training_rows = []
    for candidate in candidates:
        evaluation = full_training_evals[candidate.risk_weight]
        full_training_rows.append({
            "风险权重λ": candidate.risk_weight,
            "完整200训练池_期望利润/元": float(evaluation.total_profit.mean()),
            "完整200训练池_下尾CVaR90/元": lower_tail_cvar(evaluation.total_profit),
        })
    training_objectives = training_objectives.merge(
        pd.DataFrame(full_training_rows), on="风险权重λ", how="left"
    )
    plan_differences = plan_difference_table(candidates)
    risk_neutral_train_mean = optimization_evals[0.0].total_profit.mean()
    best_candidate_train_mean = max(ev.total_profit.mean() for ev in optimization_evals.values())
    if risk_neutral_train_mean + 1.0 < best_candidate_train_mean:
        warnings.warn(
            "风险中性候选的训练期望利润低于其他风险候选；各模型均独立且受求解时限约束，"
            "这表示 λ=0 的整数可行解尚未充分收敛，请结合 MIP 间隙解释。",
            RuntimeWarning,
        )
    if plan_differences["是否与基准方案相同"].iloc[1:].eq("是").all():
        warnings.warn(
            "所有风险权重仍得到同一方案；当前已排除跨权重固定变量，应将其解释为风险前沿退化，"
            "并结合目标差异与MIP间隙检查。",
            RuntimeWarning,
        )
    validation_evals = {c.risk_weight: evaluate_plan(c, stat_for_key, validation) for c in candidates}
    selected, frontier = choose_candidate(candidates, validation_evals)
    test_evals = {c.risk_weight: evaluate_plan(c, stat_for_key, test) for c in candidates}
    selected_test = test_evals[selected.risk_weight]
    hard_checks = validate_hard_constraints(data, selected)
    print(f"  入选 λ={selected.risk_weight:.1f}；测试期望利润={selected_test.total_profit.mean():,.0f} 元；"
          f"CVaR90={lower_tail_cvar(selected_test.total_profit):,.0f} 元")

    print("[STAGE E] 写入官方模板、分析工作簿、源数据与诊断日志……")
    write_outputs(
        data, candidates, selected, frontier, solver_diag, test_evals,
        selected_test, hard_checks, args.quick, training_objectives, plan_differences,
        scenario_reduction,
    )
    if not args.skip_figures:
        print("[STAGE F] 生成 Nature 风格中文论文图……")
        subprocess.run([sys.executable, str(SCRIPT_DIR / "Q2_nature_figures.py")], check=True)
    print(f"完成。结果目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
